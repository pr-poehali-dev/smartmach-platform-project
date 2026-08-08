"""
Генерация отчётов из JSON: HTML → PDF (WeasyPrint) и Excel (openpyxl).
Оба формата строятся из одной структуры report_schema.build_report().
"""
from __future__ import annotations

from pathlib import Path

from jinja2 import Environment, FileSystemLoader, select_autoescape

TEMPLATES_DIR = Path(__file__).parent / "templates"


def render_html(report: dict) -> str:
    """Рендер HTML-отчёта по Jinja2-шаблону."""
    env = Environment(
        loader=FileSystemLoader(TEMPLATES_DIR),
        autoescape=select_autoescape(["html"]),
    )
    return env.get_template("report.html.j2").render(r=report)


def render_pdf(report: dict, path: str | Path) -> str:
    """HTML → PDF. Требует WeasyPrint; при отсутствии сохраняет HTML."""
    html = render_html(report)
    path = Path(path)
    try:
        from weasyprint import HTML  # noqa: PLC0415
        HTML(string=html).write_pdf(str(path))
        return str(path)
    except ImportError:
        fallback = path.with_suffix(".html")
        fallback.write_text(html, encoding="utf-8")
        return str(fallback)


def render_excel(report: dict, path: str | Path) -> str:
    """Выгрузка спецификации, маршрутной карты и калькуляции в Excel."""
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import Font  # noqa: PLC0415

    wb = Workbook()
    bold = Font(bold=True)

    # Лист 1 — спецификация
    ws = wb.active
    ws.title = "Спецификация"
    ws.append(["Поз.", "Обозначение", "Наименование", "Кол.", "Материал", "Масса, кг", "Примечание"])
    for c in ws[1]:
        c.font = bold
    for i in report["specification"]["items"]:
        ws.append([i["pos"], i["designation"], i["name"], i["qty"], i["material"], i["mass_kg"], i["note"]])

    # Лист 2 — маршрутная карта
    ws2 = wb.create_sheet("Маршрутная карта")
    ws2.append(["Операция", "Наименование", "Оборудование", "Тшт, мин"])
    for c in ws2[1]:
        c.font = bold
    for op in report["route_card"]["operations"]:
        ws2.append([op["op"], op["name"], op["equipment"], op["time_min"]])
    ws2.append(["", "ИТОГО, ч", "", report["route_card"]["total_time_h"]])

    # Лист 3 — себестоимость
    ws3 = wb.create_sheet("Себестоимость")
    ws3.append(["Статья затрат", "Значение"])
    for c in ws3[1]:
        c.font = bold
    for k, v in report["cost"]["breakdown"].items():
        ws3.append([k, v])
    ws3.append(["Итого на партию, ₽", report["cost"]["cost_total"]])
    ws3.append(["На единицу, ₽", report["cost"]["cost_per_unit"]])
    ws3.append(["Срок, дней", report["cost"]["lead_days"]])

    for sheet in (ws, ws2, ws3):
        for col in sheet.columns:
            width = max(len(str(c.value or "")) for c in col) + 2
            sheet.column_dimensions[col[0].column_letter].width = min(width, 45)

    wb.save(str(path))
    return str(path)
